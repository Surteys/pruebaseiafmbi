"""
        NOTAS:

for root, dirs, files in os.walk(dir_path):
    for file in files: 
  
        # change the extension from '.mp3' to 
        # the one of your choice.
        if file.endswith('.mp3'):
            print (root+'/'+str(file))

#Para la caja TBLU se usan fusibles ATO con color claro por lo que al color se le agrega un "_clear", por ejemplo "ATO,10,red_clear"

#"F400": "ATO,15,BLUE
"""
           
from copy import copy
import requests
import openpyxl
import json
import os

modules = {}

fuses_types = {
    'PDC-P': {
        'MF1': "MULTI", 'MF2': "MULTI", 'F300': "ATO", 'F301': "MINI", 'F302': "MINI", 'F303': "MINI", 'F304': "MINI", 'F305': "MINI", 'F318': "MINI", 
        'F319': "MINI", 'F320': "MINI", 'F321': "MINI", 'F322': "MINI", 'F323': "MINI", 'F324': "MINI", 'F325': "MINI", 'F326': "ATO", 'F327': "ATO", 
        'F328': "ATO", 'F329': "ATO", 'F330': "ATO", 'F331': "ATO", 'F332': "ATO", 'F333': "ATO", 'F334': "ATO", 'F335': "ATO", 'E21': "CONN", 
        'E22': "CONN"
    },
    'PDC-D': {
        'F200': "MINI", 'F201': "MINI", 'F202': "MINI", 'F203': "MINI", 'F204': "MINI", 'F205': "MINI", 'F206': "MINI", 'F207': "MINI", 'F208': "MINI", 
        'F209': "ATO", 'F210': "ATO", 'F211': "ATO", 'F212': "ATO", 'F213': "ATO", 'F214': "ATO", 'F215': "ATO", 'F216': "ATO", 'F217': "MINI", 
        'F218': "MINI", 'F219': "MINI", 'F220': "MINI", 'F221': "MINI", 'F222': "MINI", 'F223': "MINI", 'F224': "MINI", 'F225': "MINI", 'F226': "MINI", 
        'F227': "MINI", 'F228': "MINI", 'F229': "MINI", 'F230': "MINI", 'F231': "MINI", 'F232': "MINI"
    },
    'PDC-R': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F449': "MAXI", 'F448': "MAXI", 'F447': "MAXI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 
        'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'F462': "MAXI", 
        'F463': "MAXI", 'F464': "MAXI", 'F465': "MINI", 'F466': "MINI", 'F467': "MINI", 'F468': "MINI", 'F469': "MINI", 'F470': "MINI", 'F471': "ATO", 
        'F472': "ATO", 'F473': "ATO", 'F474': "ATO", 'F475': "ATO", 'F476': "ATO", 'F477': "ATO", 'F478': "ATO", 'F479': "ATO", 'F480': "ATO", 
        'F481': "ATO", 'F482': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F96': "ATO"
    },
    'PDC-RMID': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-RS': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-S': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'TBLU': {
        '9': "ATO", '8': "ATO", '7': "ATO", '6': "ATO", '5': "ATO", '4': "ATO", '3': "ATO", '2': "ATO", '1': "ATO"
    }
}

fuses_color = {
    "1":    "black",
    "5":    "beige",
    "7.5":  "brown",
    "10":   "red",
    "15":   "blue",
    "20":   "yellow",
    "25":   "white",
    "30":   "green",
    "40":   "amber",
    "50":   "red",
    "60":   "blue"
    }

##################################### Modules management #################################
def makeModules(data):
    global modules
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://127.0.0.1:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    modules = {}
    print("#################### Modules ####################")
    print("Modulos anteriormente cargados: ",modules)
    dir_path = os.path.join(os.getcwd(), '..\\modules\\')
    file_name = None
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    if "Acomodos Modularidades" in sheet or "MFB" in sheet:
                        continue
                    currentSheet = file[sheet]
                    for column in range(8, currentSheet.max_column + 1):
                        module = currentSheet.cell(row = 3, column = column).value
                        if not(module in modules):
                            elimV = module.find("V")
                            if elimV != -1:
                                continue
                                #print("Este modulo no se agrega ala DB: ",module)
                            else:
                                modules[module] = {}
                                #print("Modulo: ", module)
                        for row in range(5,currentSheet.max_row  + 1):
                            value = currentSheet.cell(row = row, column = column).value
                            if value == "x" or value == "X":
                                box = currentSheet.cell(row = row, column = 1).value
                                if box =="Fuse Box F55":
                                    box = "TBLU"
                                fuse = currentSheet.cell(row = row, column = 2).value
                                if box == "TBLU":
                                    fuse = fuse.replace("A", "")
                                if box == "PDC-S":
                                    fuse = str(fuse)
                                    #print("Tipo del Fuse Ya convertido: ",type(fuse))
                                if box == "F96":
                                    print("Caja F96 AQUI",module)
                                    box = "PDC-RMID"
                                    print("CAJA F96 TRANSFORMADA A PDC-RMID")
                                if box == "PDC-R":
                                    if module in pdcrVariantes["large"]:
                                        box = "PDC-R"
                                    elif module in pdcrVariantes["medium"]:
                                        box = "PDC-RMID"
                                    elif module in pdcrVariantes["small"]:
                                        box = "PDC-RS"
                                    else:
                                        box = "PDC-RS"
                                    if fuse == "X" or fuse == "T" or fuse == "U":
                                        fuse = "REL" + fuse
                                amp = currentSheet.cell(row = row, column = 7).value
                                if not(box in modules[module]):
                                    modules[module][box] = {}
                                modules[module][box][fuse] = amp[:-1]
                os.remove(root+'\\'+ file_name)

    structured_data = []
    for module in modules:
        temp = {
            "DBEVENT": data,
            "MODULO": "",
            "PDC-R": {},
            "PDC-RMID": {},
            "PDC-RS": {},
            "PDC-S": {},
            "TBLU": {},
            "PDC-D": {},
            "PDC-P": {}
            }

        temp["MODULO"] = module
        for box in modules[module]:
            for fuse in modules[module][box]:
                try:
                    amp     = modules[module][box][fuse]
                    Type    = fuses_types[box][fuse]
                    color = ""
                    if Type == "RELAY":
                        if amp == "60":
                            color = "red"
                        elif amp == "70":
                            color = "gray"
                    else:
                        color   = fuses_color[amp]
                    if box == "TBLU":
                        color = color + "_clear"
                    temp[box][fuse] = Type + "," + amp + "," + color
                except Exception as ex:
                    print("\nexception in [", module, "] [", box, "] [", fuse, "]")
                    print(ex)
        structured_data.append(temp)

    print ("\n total de modulos: ",len(structured_data))

    return structured_data

def updateModules(data):
    print("updating")
    tabla = data[0]["DBEVENT"]
    print("TABLAAAAA Update Modules+-+-+-+-: ",tabla)
    endpoint = f"http://127.0.0.1:5000/api/get/{tabla}/modulos_fusibles/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = "http://127.0.0.1:5000/api/post/modulos_fusibles"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://127.0.0.1:5000/api/update/modulos_fusibles/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def pdcrVariants (data):
    """

            IN CONSTRUCTION


    PDC-R small:  A2239060902
    PDC-R MEDIUM:  A2239061002
    PDC-R LARGE:  A2239061102
    """
    print("#################### pdcrVariants ####################")
    dir_path = os.path.join(os.getcwd(), '..\\FAAJISPREV\\')
    file_name = None
    rows = []
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if temp.endswith('.txt'):
                fic = open(dir_path + file_name)
                lines = list(fic)
                for i in lines:
                    i = i[:-1]
                    rows.append(i.split())
                print(len(lines))
                for i in range(5):
                    print(lines[i])

def refreshModules(data):
    data = makeModules(data)
    updateModules(data)

################################### Modularities management ##############################
def makeModularities(data):
    global modules
    print("#################### Modularities ####################")
    endpoint = f"http://127.0.0.1:5000/api/get/{data}/modulos_fusibles/all/-/-/-/-/-"
    modulesExisting = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos: ",modulesExisting["MODULO"])
    dir_path = os.path.join(os.getcwd(), '..\\ILX\\')
    file_name = None
    modularities = []
    modulosFaltantes = []
    ilxfaltantes = {
        "ILX": {},
        "Modulos": {}
        }
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if temp.endswith('.dat'):
                fic = open(dir_path + file_name)
                lines = list(fic)
                csv = ""
                for line in lines:
                    csv += line.rsplit(sep = "=")[-1][:-1] + ","
                csv = csv[:-1]
                fic.close()
                temp = {
                    "DBEVENT": data,
                    "MODULARIDAD": ILX,
                    "FECHA": "AUTO",
                    "MODULOS_FUSIBLES": csv,
                    "ACTIVO": 1
                    }
                #print("ILX: ",ILX)
                #print("Modulos que tiene: ",csv)
                #print("Modulos que tiene TIPO: ",type(csv))
                #print("Modulos que tiene el ILX: ",csv.split(","))
                #print("Modulos que tiene convertido a array TIPO: ",type(csv.split(",")))
                modulosDesconocidos = set(csv.split(",")) - set(modulesExisting["MODULO"])
                #print("Comparación; Modulos del ILX que NO están en la base de datos: ", modulosDesconocidos)
                #print("Comparación; Modulos del ILX que NO están en la base de datos LEN: ", len(modulosDesconocidos))
                #print("Comparación tipo", type(modulosDesconocidos))
                if len(modulosDesconocidos) == 0:
                    modularities.append(temp)
                else:
                    ilxfaltantes["ILX"][ILX] = []
                    for e in modulosDesconocidos:
                        ilxfaltantes["ILX"][ILX].append(e)
                    #print(e)
                        if not(e in modulosFaltantes):
                            modulosFaltantes.append(e)
                    ilxfaltantes["Modulos"] = modulosFaltantes
                os.remove(root+'\\'+ file_name)

        #print("MODULOS FALTANTES FINAL : ",modulosFaltantes)
        #print("ILX que NO se cargaron a la estación : ",ilxfaltantes)
    if len(modularities) != 0:
        updateModularities(modularities)
    return ilxfaltantes
 
def updateModularities(data):
    print("updating")
    print("Data dentro de Upload Modularities: ",data)
    tabla = data[0]["DBEVENT"]
    endpoint = f"http://127.0.0.1:5000/api/get/{tabla}/modularidades/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULARIDAD" in existing):
        existing["MODULARIDAD"] = []
    for i in data:
        try:
            if not(i["MODULARIDAD"] in existing["MODULARIDAD"]):
                endpoint = "http://127.0.0.1:5000/api/post/modularidades"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULARIDAD"].index(i["MODULARIDAD"])
                id = existing["ID"][index]
                endpoint = f"http://127.0.0.1:5000/api/update/modularidades/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)


if __name__ == '__main__':
    print("finished")
    #refreshModules()
    #data = makeModules()
    #updateModules(data)
    #pdcrVariants("dumie")